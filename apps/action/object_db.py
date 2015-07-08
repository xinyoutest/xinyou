# coding:utf-8

import selenium
import os
import time
import sys
reload(sys)
sys.setdefaultencoding("utf-8")
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

# 全局目录地址
action_path = os.path.dirname(__file__)
apps_path = os.path.dirname(action_path)
model_path = apps_path + '\\model'
report_path = apps_path + '\\report'
image_path = report_path + '\\image'
data_path = apps_path + '\\data'
xinyou_path = os.path.dirname(apps_path)

# 初始化chrome浏览器方法
def chrome():
    options = webdriver.ChromeOptions()
    options.add_experimental_option(
        "excludeSwitches", ["ignore-certificate-errors"])
    return webdriver.Chrome(chrome_options=options)

# 初始化firefox浏览器方法
def firefox():
    return webdriver.Firefox()


# 打开初始化页面并最大化浏览器
def start_Browser(self):
    self.dr.get(self.url)
    self.dr.maximize_window()


# 设置智能等待时间
def set_implicitly_wait(self, time):
    self.dr.implicitly_wait(time)

# 设置页面加载等待时间
def set_page_toload_timeout(self, time):
    self.dr.set_page_load_timeout(time)


# openpyxl读取execl文件
def getExecl(sheet, c, r):
    '''读取测试用例文件，返回指定sheet、列、行的数据'''
    file_name = data_path + '\\auto_case.xlsx'
    # 获取一个已经存在的excel文件wb
    wb = load_workbook(file_name)
    ws = wb.get_sheet_by_name(sheet)
    values = ws.cell(column=c, row=r).value
    return values


def writeExcel(sheet, r, result):
    '''写入测试用例文件内容，指定sheet、列、行的数据'''
    file_name = data_path + '\\auto_case.xlsx'
    # 获取一个已经存在的excel文件wb
    wb = load_workbook(file_name)
    ws = wb.get_sheet_by_name(sheet)
    ws.cell(column=11, row=r).value = result  # 执行,N/A#未执行
    now = time.strftime("%Y%m%d.%H%M%S", time.localtime(time.time()))
    ws.cell(column=12, row=r).value = now
    wb.save(file_name)


def findId(driver, id):
    return driver.find_element_by_id(id)

# 保存图片，存放到指定的文件夹
def save_error_image(self):
    #@xu
    # now = time.strftime("%Y%m%d.%H%M%S", time.localtime(time.time()))
    now = time.strftime("%Y%m%d")
    file_name = image_path + '\\' + self.error_image_name + "-" + now + '.png'
    self.dr.get_screenshot_as_file(file_name)

def getCookie(self, cookie_name):
    # cookies=self.dr.get_cookies()
    # for cookie in cookies:
    #     if cookie == cookie_name:
    #         cookie_value = cookie["Value"]
    #         return cookie_value
    cookie = self.dr.get_cookie(cookie_name)
    return cookie["value"].replace("%40", "@")